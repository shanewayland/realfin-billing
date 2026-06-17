from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from datetime import datetime, timedelta
import io

app = Flask(__name__)
CORS(app)

currency_fmt = '_("$"* #,##0.00_);_("$"* \\(#,##0.00\\);_("$"* "-"??_);_(@_)'
pct_fmt = '0.00%'

def parse_date(s):
    if not s:
        return datetime.now()
    for fmt in ('%m/%d/%Y', '%m/%d/%y', '%Y-%m-%d'):
        try:
            return datetime.strptime(str(s), fmt)
        except:
            pass
    return datetime.now()

def set_cell(ws, coord, value, bold=False, align=None, number_format=None):
    cell = ws[coord]
    cell.value = value
    cell.font = Font(name='Aptos Narrow', size=11, bold=bold)
    if align:
        cell.alignment = Alignment(horizontal=align, wrap_text=(align == 'wrap'))
    if number_format:
        cell.number_format = number_format

@app.route('/generate', methods=['POST'])
def generate():
    data = request.json
    loan = data.get('loan', {})
    activities = data.get('activities', [])

    activities = [a for a in activities if a.get('d') and a.get('t') not in ('EDPC', 'Notes')]
    activities.sort(key=lambda x: parse_date(x.get('d')))

    if activities:
        first_date = parse_date(activities[0]['d'])
    else:
        first_date = parse_date(loan.get('fd', '01/01/2026'))

    billing_month_start = first_date.replace(day=1)
    if billing_month_start.month == 12:
        next_month = billing_month_start.replace(year=billing_month_start.year + 1, month=1)
    else:
        next_month = billing_month_start.replace(month=billing_month_start.month + 1)
    billing_month_end = next_month - timedelta(days=1)
    statement_date = next_month

    running_balance = float(loan.get('bp') or loan.get('bal') or 0)
    current_rate = float(loan.get('rate') or 0)
    loan_spread = float(loan.get('spread') or 0)
    floor_rate = float(loan.get('floor') or 0)

    rows = []
    total_interest = 0
    segment_start = billing_month_start

    rows.append({
        'memo': 'Balance Forward',
        'type': '',
        'principal': running_balance,
        'trans': 0,
        'dates': f"{billing_month_start.strftime('%m/%d/%Y')} - {billing_month_start.strftime('%m/%d/%Y')}",
        'days': 0,
        'rate': current_rate if current_rate else None,
        'interest': 0
    })

    for act in activities:
        act_date = parse_date(act['d'])
        days = (act_date - segment_start).days + 1

        activity_type = act.get('t', '')
        trans_amt = 0

        dis = float(act.get('dis') or 0)
        pp = float(act.get('pp') or 0)
        ip = float(act.get('ip') or 0)
        pr = act.get('pr')

        if dis:
            running_balance += dis
            trans_amt = dis
        if pp:
            running_balance -= pp
            trans_amt = pp
        if ip:
            running_balance += ip
            trans_amt = ip
        if pr is not None and pr != '':
            new_prime = float(pr)
            current_rate = max(loan_spread + new_prime, floor_rate)
            trans_amt = 0

        interest = round(running_balance * current_rate / 360 * days, 2) if days > 0 else 0
        total_interest += interest

        rows.append({
            'memo': activity_type,
            'type': '',
            'principal': running_balance,
            'trans': trans_amt,
            'dates': f"{segment_start.strftime('%m/%d/%Y')} - {act_date.strftime('%m/%d/%Y')}",
            'days': days,
            'rate': current_rate,
            'interest': interest
        })

        segment_start = act_date + timedelta(days=1)

    if activities:
        last_act_date = parse_date(activities[-1]['d'])
        final_days = (billing_month_end - last_act_date).days
        if final_days > 0:
            last_row = rows[-1]
            last_row['dates'] = f"{last_act_date.strftime('%m/%d/%Y')} - {billing_month_end.strftime('%m/%d/%Y')}"
            last_row['days'] = final_days + 1
            last_row['interest'] = round(running_balance * current_rate / 360 * last_row['days'], 2)
            total_interest = round(sum(r['interest'] for r in rows[:-1]) + last_row['interest'], 2)

    total_interest = round(total_interest, 2)
    all_rows = rows

    wb = Workbook()
    ws = wb.active
    ws.title = 'Billing Statement'

    col_widths = {
        'A': 35, 'B': 20, 'C': 18, 'D': 20, 'E': 28, 'F': 12, 'G': 30, 'H': 16
    }
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width

    ws.merge_cells('A1:H1')
    set_cell(ws, 'A1', 'LOAN BILLING STATEMENT', bold=True, align='center')

    set_cell(ws, 'A2', loan.get('bn', ''))
    set_cell(ws, 'G2', 'As of Date:', align='right')
    set_cell(ws, 'H2', billing_month_end.strftime('%m/%d/%Y'))

    set_cell(ws, 'A4', '1111 North Post Oak Road')
    set_cell(ws, 'G4', 'Statement Date:', align='right')
    set_cell(ws, 'H4', statement_date.strftime('%m/%d/%Y'))

    set_cell(ws, 'A5', 'Houston, Texas 77055')

    set_cell(ws, 'A9', 'Loan Number / Unit:', align='right')
    set_cell(ws, 'B9', loan.get('ln', ''))

    set_cell(ws, 'A10', 'Address:', align='right')
    set_cell(ws, 'B10', loan.get('pa', ''))

    set_cell(ws, 'B13', 'Loan Commitment:', bold=True)
    set_cell(ws, 'C13', float(loan.get('na') or 0), bold=True, number_format=currency_fmt)

    set_cell(ws, 'A15', 'Memo Description', bold=True)
    set_cell(ws, 'C15', 'Billing Date', bold=True)
    set_cell(ws, 'D15', 'Due Date', bold=True)
    set_cell(ws, 'E15', 'Amount Due', bold=True)

    set_cell(ws, 'A16', 'INTEREST BILLING - PERIOD END')
    set_cell(ws, 'C16', billing_month_end.strftime('%m/%d/%Y'), align='left')
    set_cell(ws, 'D16', statement_date.strftime('%m/%d/%Y'), align='left')
    set_cell(ws, 'E16', total_interest, bold=True, number_format=currency_fmt)

    set_cell(ws, 'D17', 'Total:')
    set_cell(ws, 'E17', total_interest, bold=True, number_format=currency_fmt)

    headers = [('A','Memo Description'),('C','Principal Balance'),
               ('D','Transaction Amount'),('E','From / To Date'),('F','# of Days'),
               ('G','Rate'),('H','Interest Due')]
    for col, val in headers:
        set_cell(ws, f'{col}20', val, bold=True, align='center')

    for i, row in enumerate(all_rows):
        r = 21 + i
        set_cell(ws, f'A{r}', row['memo'])
        set_cell(ws, f'B{r}', row['type'])
        set_cell(ws, f'C{r}', row['principal'], number_format=currency_fmt)
        set_cell(ws, f'D{r}', row['trans'], number_format=currency_fmt)
        set_cell(ws, f'E{r}', row['dates'])
        set_cell(ws, f'F{r}', row['days'])
        if row['rate'] is not None:
            set_cell(ws, f'G{r}', row['rate'], number_format=pct_fmt)
        set_cell(ws, f'H{r}', row['interest'], number_format=currency_fmt)

    total_row = 21 + len(all_rows)
    set_cell(ws, f'B{total_row}', 'Total:', bold=True, align='right')
    set_cell(ws, f'C{total_row}', running_balance, bold=True, number_format=currency_fmt)
    set_cell(ws, f'D{total_row}', sum(r['trans'] for r in all_rows), bold=True, number_format=currency_fmt)
    set_cell(ws, f'G{total_row}', 'Total Interest for the Month:', bold=True, align='right')
    set_cell(ws, f'H{total_row}', total_interest, bold=True, number_format=currency_fmt)

    pay_row = total_row + 3
    set_cell(ws, f'G{pay_row}', 'PLEASE PAY THIS AMOUNT:', bold=True, align='right')
    set_cell(ws, f'H{pay_row}', total_interest, bold=True, number_format=currency_fmt)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"{loan.get('ln', 'Loan')}_Billing_Statement.xlsx"
    return send_file(buffer, as_attachment=True, download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/health', methods=['GET'])
def health():
    return jsonify({'status': 'ok'})

if __name__ == '__main__':
    app.run(debug=True)
